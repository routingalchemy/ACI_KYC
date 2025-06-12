#!/usr/bin/env python

import json
import sys
import requests
import argparse
from openpyxl import Workbook
from openpyxl import load_workbook
from urllib3.exceptions import InsecureRequestWarning

requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)

__author__ = "Richard Buknicz"
__copyright__ = "Copyright 2024, Routingalchemy ACI_KYC Project"
__license__ = "GPL"
__version__ = "0.4"
__maintainer__ = "Richard Buknicz"
__status__ = "Production"


class aci_kyc:
    """ACI Know Your Contracts. A Contract exporter and visualiser tool"""

    def __init__(self, hostname, username, password):
        """Host resource definition"""
        self.hostname = hostname
        self.username = username
        self.password = password
        self.token = {}
        self.url = ""
        self.http_headers = {
            "Content-Type": "application/json",
            "Accept": "application/json",
        }
        self.http_cert_verify = False
        self.apic_token()

    def apic_token(self):
        """Getting a token from the ACI APIC"""
        host = f"https://{self.hostname}/api/aaaLogin.json"
        data = {"aaaUser": {"attributes": {"name": self.username, "pwd": self.password}}}
        cookie_key = "APIC-cookie"
        cookie_value = 'response.json()["imdata"][0]["aaaLogin"]["attributes"]["token"]'

        try:

            response = requests.post(
                host,
                headers=self.http_headers,
                data=json.dumps(data),
                verify=self.http_cert_verify,
            )
            response.raise_for_status()
            if response.status_code != 204:
                self.token = {cookie_key: eval(cookie_value)}
        except requests.exceptions.RequestException as error:
            raise SystemExit(error)

    def __api_get(self):
        """Get request for retrieving the contract data"""

        host = f"https://{self.hostname}{self.url}"
        try:
            response = requests.get(
                host,
                cookies=self.token,
                headers=self.http_headers,
                verify=self.http_cert_verify,
            )
            response.raise_for_status()  # raises exception when not a 2xx response
            return response.json()
        except requests.exceptions.RequestException as error:
            raise SystemExit(error)

    def __contract_details(self, contract_dn):
        """Get a contract detail (DN)"""
        self.url = f"/api/node/mo/{contract_dn}.json?query-target=subtree"
        brcp_dn_data = self.__api_get()
        source_grp, destination, subject, filter = [], [], [], []
        for brcp_dn_imdata in brcp_dn_data["imdata"]:
            if "vzRtCons" in brcp_dn_imdata.keys():  # consumer/source details
                source_grp.append(
                    {
                        "name": brcp_dn_imdata["vzRtCons"]["attributes"]["tDn"].split("/")[-1][4:],
                        "app": brcp_dn_imdata["vzRtCons"]["attributes"]["dn"].split("/")[-2][3:],
                        "tenant": brcp_dn_imdata["vzRtCons"]["attributes"]["dn"].split("/")[1][3:],
                        "type": self.__object_norm(brcp_dn_imdata["vzRtCons"]["attributes"]["tCl"]),
                    }
                )
            if "vzRtProv" in brcp_dn_imdata.keys():  # provider/destination details
                destination.append(
                    {
                        "name": brcp_dn_imdata["vzRtProv"]["attributes"]["tDn"].split("/")[-1][4:],
                        "app": brcp_dn_imdata["vzRtProv"]["attributes"]["dn"].split("/")[-2][3:],
                        "tenant": brcp_dn_imdata["vzRtProv"]["attributes"]["dn"].split("/")[1][3:],
                        "type": self.__object_norm(brcp_dn_imdata["vzRtProv"]["attributes"]["tCl"]),
                    }
                )
            if "vzRtAnyToCons" in brcp_dn_imdata.keys():  # vzany sources
                source_grp.append(
                    {
                        "name": brcp_dn_imdata["vzRtAnyToCons"]["attributes"]["tDn"].split("/")[-2][
                            4:
                        ],
                        "app": "",
                        "tenant": brcp_dn_imdata["vzRtAnyToCons"]["attributes"]["dn"].split("/")[1][
                            3:
                        ],
                        "type": self.__object_norm(
                            brcp_dn_imdata["vzRtAnyToCons"]["attributes"]["tCl"]
                        ),
                    }
                )
            if "vzRtAnyToProv" in brcp_dn_imdata.keys():  # vzany providers
                destination.append(
                    {
                        "name": brcp_dn_imdata["vzRtAnyToProv"]["attributes"]["tDn"].split("/")[-2][
                            4:
                        ],
                        "app": "",
                        "tenant": brcp_dn_imdata["vzRtAnyToProv"]["attributes"]["dn"].split("/")[1][
                            3:
                        ],
                        "type": self.__object_norm(
                            brcp_dn_imdata["vzRtAnyToProv"]["attributes"]["tCl"]
                        ),
                    }
                )
            if "vzSubj" in brcp_dn_imdata.keys():  # subjects
                subject.append(
                    {
                        "name": brcp_dn_imdata["vzSubj"]["attributes"]["dn"].split("/")[-1][5:],
                        "revfltports": brcp_dn_imdata["vzSubj"]["attributes"]["revFltPorts"],
                    }
                )
                self.url = f"/api/node/mo/{brcp_dn_imdata['vzSubj']['attributes']['dn']}.json?query-target=subtree"
                subject_dn_data = self.__api_get()
                for subject_dn_imdata in subject_dn_data["imdata"]:
                    sgraph = "N/A"
                    if "vzRsSubjGraphAtt" in subject_dn_imdata.keys():
                        sgraph = subject_dn_imdata["vzRsSubjGraphAtt"]["attributes"][
                            "tnVnsAbsGraphName"
                        ]
                    if ("vzRsSubjFiltAtt" or "vzRsFiltAtt") in subject_dn_imdata.keys():
                        fltatt = next(iter(subject_dn_imdata.keys()))
                        filter.append({"action": subject_dn_imdata[fltatt]["attributes"]["action"]})
                        if subject_dn_imdata[fltatt]["attributes"]["tDn"] != "":
                            self.url = f"/api/node/mo/{subject_dn_imdata[fltatt]['attributes']['tDn']}.json?query-target=subtree"
                            filter_dn_data = self.__api_get()
                            entries = []
                            for filter_dn_imdata in filter_dn_data["imdata"]:
                                if "vzFilter" in filter_dn_imdata.keys():
                                    filter[len(filter) - 1]["name"] = filter_dn_imdata["vzFilter"][
                                        "attributes"
                                    ]["name"]
                                if "vzEntry" in filter_dn_imdata.keys():
                                    entries.append(
                                        {
                                            "name": filter_dn_imdata["vzEntry"]["attributes"][
                                                "dn"
                                            ].split("/")[-1][2:],
                                            "etht": filter_dn_imdata["vzEntry"]["attributes"][
                                                "etherT"
                                            ],
                                            "sport": self.__port_compare(
                                                filter_dn_imdata["vzEntry"]["attributes"][
                                                    "sFromPort"
                                                ],
                                                filter_dn_imdata["vzEntry"]["attributes"][
                                                    "sToPort"
                                                ],
                                            ),
                                            "dport": self.__port_compare(
                                                filter_dn_imdata["vzEntry"]["attributes"][
                                                    "dFromPort"
                                                ],
                                                filter_dn_imdata["vzEntry"]["attributes"][
                                                    "dToPort"
                                                ],
                                            ),
                                            "stateful": filter_dn_imdata["vzEntry"]["attributes"][
                                                "stateful"
                                            ],
                                            "tcprules": filter_dn_imdata["vzEntry"]["attributes"][
                                                "tcpRules"
                                            ],
                                            "icmp": f"icmpv4: {self.__object_norm(filter_dn_imdata['vzEntry']['attributes']['icmpv4T'])} \n icmpv6: {self.__object_norm(filter_dn_imdata['vzEntry']['attributes']['icmpv6T'])}",
                                            "applyToFrag": filter_dn_imdata["vzEntry"][
                                                "attributes"
                                            ]["applyToFrag"],
                                        }
                                    )
                            filter[len(filter) - 1]["entries"] = entries
                            subject[len(subject) - 1]["filter"] = filter
                        subject[len(subject) - 1]["sgraph"] = sgraph
        return source_grp, destination, subject

    def contract_info(self, **kwargs):
        """Get contract info from the fabric"""
        contract_list = []
        self.url = '/api/node/class/vzBrCP.json?query-target-filter=and(wcard(vzBrCP.dn,"{}"),wcard(vzBrCP.name,"{}"))'.format(
            kwargs["dn"], kwargs["contract"]
        )
        brcp_data = self.__api_get()
        if int(brcp_data["totalCount"]) == 0:
            sys.exit("0 contaracts found. Contract or Tenant name not defined properly")
        print(f"{brcp_data['totalCount']} contracts found")
        for brcp_imdata in brcp_data["imdata"]:
            cname = brcp_imdata["vzBrCP"]["attributes"]["name"]
            ctenant = brcp_imdata["vzBrCP"]["attributes"]["dn"].split("/")[1].split("tn-")[1]
            print(f"Extracting info for {cname} contract from {ctenant} tenant")
            contract_list.append(
                {
                    "name": cname,
                    "tenant": ctenant,
                    "scope": brcp_imdata["vzBrCP"]["attributes"]["scope"],
                }
            )
            source_grp, destination, subject = self.__contract_details(
                brcp_imdata["vzBrCP"]["attributes"]["dn"]
            )
            contract_list[len(contract_list) - 1]["source"] = source_grp
            contract_list[len(contract_list) - 1]["destination"] = destination
            contract_list[len(contract_list) - 1]["subject"] = subject
        return contract_list

    def contract2excel(self, clist):
        """Output the contract data to excel"""
        wb = load_workbook(filename="contracts_template.xlsx")
        row_offset = 3
        for contract in clist:
            ws = wb.copy_worksheet(wb["template"])
            ws.title = f"{contract['name']}"
            ws["B1"] = contract["name"]
            ws["D1"] = contract["tenant"]
            ws["F1"] = contract["scope"]
            for soi in range(len(contract["source"])):  # source and destination for loop to colapse
                ws[f"A{soi + row_offset}"] = (
                    f"{contract['source'][soi]['tenant']}:{contract['source'][soi]['app']}:{contract['source'][soi]['name']}"
                )
                ws[f"B{soi + row_offset}"] = f"{contract['source'][soi]['type']}"
            for dei in range(len(contract["destination"])):
                ws[f"M{dei + row_offset}"] = (
                    f"{contract['destination'][dei]['tenant']}:{contract['destination'][dei]['app']}:{contract['destination'][dei]['name']}"
                )
                ws[f"N{dei + row_offset}"] = f"{contract['destination'][dei]['type']}"
            entry_row_offset = 3
            for subject_it in contract["subject"]:
                ws[f"C{entry_row_offset}"] = subject_it["name"]
                ws[f"L{entry_row_offset}"] = subject_it["sgraph"]
                smfrom = entry_row_offset
                if "filter" in subject_it:
                    for filter_it in subject_it["filter"]:
                        fmfrom = entry_row_offset
                        ws[f"D{entry_row_offset}"] = filter_it["name"]
                        ws[f"E{entry_row_offset}"] = filter_it["action"]
                        entry_size = len(filter_it["entries"])
                        for eni in range(entry_size):
                            ws[f"F{eni + entry_row_offset}"] = filter_it["entries"][eni]["name"]
                            ws[f"G{eni + entry_row_offset}"] = filter_it["entries"][eni]["etht"]
                            ws[f"H{eni + entry_row_offset}"] = filter_it["entries"][eni]["sport"]
                            ws[f"I{eni + entry_row_offset}"] = filter_it["entries"][eni]["dport"]
                            ws[f"J{eni + entry_row_offset}"] = filter_it["entries"][eni]["stateful"]
                            ws[f"K{eni + entry_row_offset}"] = filter_it["entries"][eni]["tcprules"]
                            ws.merge_cells(f"D{fmfrom}:D{fmfrom + entry_size - 1}")
                            ws.merge_cells(f"E{fmfrom}:E{fmfrom + entry_size - 1}")
                        entry_row_offset += entry_size
                        ws.merge_cells(f"C{smfrom}:C{entry_row_offset - 1}")
                        ws.merge_cells(f"L{smfrom}:L{entry_row_offset - 1}")
        wb.remove(wb["template"])
        wb.save("contracts.xlsx")

    def __port_compare(self, fport, tport):
        """Auxaliry function for simplifying the port representation"""
        if fport == tport:
            return self.__object_norm(fport)
        else:
            return f"{self.__object_norm(fport)}-{self.__object_norm(tport)}"

    def __object_norm(self, changeme):
        """Auxaliry function to replace ACI object names with meaninful ones"""
        match changeme:
            case "unspecified":
                return "any"
            case "fvAEPg":
                return "EPG"
            case "fvESg":
                return "ESG"
            case "l2extInstP":
                return "L2Out"
            case "l3extInstP":
                return "L3Out"
            case _:
                return changeme


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="ACI - Know Your Contracts. Visualising ACI contract"
    )
    parser.add_argument("-u", "--user", help="Username", required=True, metavar="username")
    parser.add_argument("-p", "--passwd", help="Password", required=True, metavar="password")
    parser.add_argument(
        "-a", "--apic", help="APIC IP/URL", required=True, metavar="apic_ip_or_fqdn"
    )
    parser.add_argument(
        "-c",
        "--contract",
        help="Contract search/match ",
        nargs="?",
        default="",
        metavar="contract",
    )
    parser.add_argument(
        "-d",
        "--dn",
        help="Distinguished name seach/match. Alows to match a Tenant or Contract ",
        nargs="?",
        default="",
        metavar="dn",
    )
    args = parser.parse_args()

    kyc_data = aci_kyc(args.apic, args.user, args.passwd)

    kyc_list = kyc_data.contract_info(dn=args.dn, contract=args.contract)  # specific contract

    kyc_data.contract2excel(kyc_list)
